"""
This script reads an Excel file called "Data for redistricting.xlsx" that has a column called "GEOID" holding IDs of shapefiles and a column called "district" holding ID of district on a sheet called "Tractlevel".
Districts are numbered starting with 1.
Further the script reads an Excel file called "Adjacency matrix.xlsx" with two clomuns called "GEOID" on a sheet called "Adjacency".
The Excel files are located in the same folder as the script.
The script returns, for each district, lists holding contiguous GEOIDs hubs of that district.
If only on list is returned for all districts, the map is contiguous
The lists are saved in a file "contiguityInformation.txt". The file is created by this script.
"""

import numpy as np
import openpyxl
import copy

class CensusTract:
    def __init__(self, iD, neighboors, district):
        self.iD = iD  
        self.neighboors = neighboors
        self.district = district

class District:
    def __init__(self, iD, elements): 
        self.iD = iD  
        self.elements = elements  
    
def getCensusTracts(sheetInfo, sheetAdjacency):
    censusTractDictionary = {}
    for i in range(2, sheetInfo.max_row+1):
        GEOID = sheetInfo['B'+str(i)].value
        district = sheetInfo['C'+str(i)].value - 1
        neighboors = []
        for j in range(2, sheetAdjacency.max_row+1):
            if int(sheetAdjacency['A'+str(j)].value) == GEOID:
                GEOIDNeighboor = sheetAdjacency['B'+str(j)].value
                for k in range(2, sheetInfo.max_row+1):
                    if str(sheetInfo['B'+str(k)].value) == GEOIDNeighboor:
                        neighboors.append(sheetInfo['B'+str(k)].value)
        censusTractDictionary[GEOID] = CensusTract(GEOID, neighboors, district)
    return censusTractDictionary

def getDistrict(censusTractDictionary):
    districtDictionary = {}
    districtCount = 0
    for key in censusTractDictionary:
        if censusTractDictionary[key].district > districtCount:
            districtCount = censusTractDictionary[key].district
    for i in range(districtCount+1):
        districtDictionary[i] = District(i, [])
    for key in censusTractDictionary:
        districtDictionary[censusTractDictionary[key].district].elements.append(key)
    return districtDictionary

def checkIfDistrictcontiguous(districtDictionary, censusTractDictionary, key):
    contiguousElements = copy.deepcopy(districtDictionary[key].elements)
    toFill = [[contiguousElements[0]]] 
    part = 0
    del contiguousElements[0]
    while len(contiguousElements) > 0:
        toDelete = []
        lenBefore = len(contiguousElements)        
        for i in range(len(contiguousElements)):
            if i <= len(contiguousElements)-1:
                for j in range(len(toFill[part])):
                    if contiguousElements[i] in censusTractDictionary[toFill[part][j]].neighboors and contiguousElements[i] not in toFill[part] and i not in toDelete:
                        toDelete.append(i)                        
        toDelete.sort()
        toDelete.reverse()
        for i in range(len(toDelete)):
            toFill[part].append(contiguousElements[toDelete[i]])
            del contiguousElements[toDelete[i]]
        lenAfter = len(contiguousElements)      
        if lenBefore == lenAfter:
            part += 1
            toFill.append([contiguousElements[0]])
        contiguousElements = copy.deepcopy(contiguousElements)
    return toFill

def checkMapContiguity(districtDictionary, censusTractDictionary):
    for key in districtDictionary:   
        print('key:', key)        
        districtContiguity = checkIfDistrictcontiguous(districtDictionary, censusTractDictionary, key)
        writefile = open('contiguityInformation.txt', 'a')
        keyString = str(key+1)
        writefile.write('contiguous elements of district ')
        writefile.write(keyString)
        writefile.write(': ')
        writefile.write('\n')
        for i in range(len(districtContiguity)):
            districtContiguityString = str(districtContiguity[i])
            writefile.write(districtContiguityString)
            writefile.write('\n')
        writefile.write('\n')
        writefile.close()
        
CensusTractInfo = openpyxl.load_workbook('Data for redistricting.xlsx')
AdjacencyInfo = openpyxl.load_workbook('Adjacency matrix.xlsx')
sheetAdjacency = AdjacencyInfo.get_sheet_by_name('Adjacency')
sheetInfo = CensusTractInfo.get_sheet_by_name('Tractlevel')

censusTractDictionary = getCensusTracts(sheetInfo, sheetAdjacency)
print('1')  
districtDictionary = getDistrict(censusTractDictionary)
print('2')  
contiguity = checkMapContiguity(districtDictionary, censusTractDictionary)
print('3')  
